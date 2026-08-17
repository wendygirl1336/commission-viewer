from __future__ import annotations

import json
import mimetypes
import hmac
import os
import re
import secrets
import socket
import struct
import time
from email import policy
from email.parser import BytesParser
from http import cookies
from http.server import BaseHTTPRequestHandler, ThreadingHTTPServer
from io import BytesIO
from pathlib import Path
from typing import Any

from openpyxl import load_workbook


ROOT = Path(__file__).resolve().parent
DATA_FILE = ROOT / "current_upload.json"
ADMIN_ID = "admin"
ADMIN_PASSWORD = "0716"
VIEWER_ID = "company"
VIEWER_PASSWORD = "1234"
ADMIN_LOGIN_IDS = {ADMIN_ID, "관리자"}
VIEWER_LOGIN_IDS = {VIEWER_ID, "회사"}
SESSION_SECONDS = 60 * 60 * 4
MAX_LOGIN_FAILURES = 5
LOCK_SECONDS = 60 * 5
MAX_UPLOAD_BYTES = 20 * 1024 * 1024
MAX_JSON_BYTES = 64 * 1024
MAX_PARSE_COLS = 220
SESSIONS: dict[str, dict[str, Any]] = {}
LOGIN_FAILURES: dict[str, dict[str, Any]] = {}


def empty_upload(message: str = "", error: str = "") -> dict[str, Any]:
    return {"rows": [], "fileName": "", "message": message, "error": error}


def normalize_row(row: dict[str, Any]) -> dict[str, Any]:
    def saved_num(value: Any) -> float:
        try:
            if value is None or value == "":
                return 0.0
            return float(str(value).replace(",", ""))
        except Exception:
            return 0.0

    item = dict(row)
    item["company"] = clean(item.get("company", ""))
    item["product"] = clean(item.get("product", ""))
    item["source"] = clean(item.get("source", ""))
    item["year1"] = saved_num(item.get("year1", 0))
    item["year2"] = saved_num(item.get("year2", 0))
    item["year3"] = saved_num(item.get("year3", 0))
    item["year4"] = saved_num(item.get("year4", 0))
    item["insuranceType"] = item.get("insuranceType") or "life"
    item["metricMode"] = item.get("metricMode") or "percent"
    saved_total = saved_num(item.get("total", 0))
    if item["insuranceType"] == "nonlife" and item["metricMode"] == "amount":
        for key in ("year1", "year2", "year3", "year4"):
            item[key] *= 100
        saved_total *= 100
        item["metricMode"] = "percent"
    if item["insuranceType"] == "nonlife":
        item["total"] = saved_total if saved_total else item["year1"] + item["year2"] + item["year3"] + item["year4"]
        return item
    item["total"] = saved_total if item["metricMode"] == "amount" and saved_total else item["year1"] + item["year2"] + item["year3"] + item["year4"]
    return item


def load_saved_upload() -> dict[str, Any]:
    if not DATA_FILE.exists():
        return empty_upload()
    try:
        data = json.loads(DATA_FILE.read_text(encoding="utf-8"))
        if isinstance(data, dict) and isinstance(data.get("rows"), list):
            return {
                "rows": [
                    normalize_rate_displays(normalize_row(row))
                    for row in data.get("rows", [])
                    if isinstance(row, dict) and not row_is_header_noise(row) and not row_is_bad_saved_row(row)
                ],
                "fileName": data.get("fileName", ""),
                "message": data.get("message", ""),
                "error": data.get("error", ""),
            }
    except Exception:
        pass
    return empty_upload()


def save_upload(data: dict[str, Any]) -> None:
    DATA_FILE.write_text(json.dumps(data, ensure_ascii=False), encoding="utf-8")


def merge_uploaded_rows(existing_rows: list[dict[str, Any]], new_rows: list[dict[str, Any]]) -> list[dict[str, Any]]:
    existing_rows = [row for row in existing_rows if not row_is_header_noise(row) and not row_is_bad_saved_row(row)]
    new_rows = [row for row in new_rows if not row_is_header_noise(row)]
    new_sources = {
        (str(row.get("insuranceType", "life")), str(row.get("source", "")))
        for row in new_rows
        if str(row.get("source", ""))
    }
    new_companies = {
        (str(row.get("insuranceType", "life")), str(row.get("company", "")))
        for row in new_rows
        if str(row.get("company", ""))
    }
    if not new_sources:
        return existing_rows + new_rows
    kept_rows = [
        row
        for row in existing_rows
        if (str(row.get("insuranceType", "life")), str(row.get("source", ""))) not in new_sources
        and (str(row.get("insuranceType", "life")), str(row.get("company", ""))) not in new_companies
    ]
    return kept_rows + new_rows


def row_is_bad_saved_row(row: dict[str, Any]) -> bool:
    if str(row.get("insuranceType", "life")) != "life":
        return False
    company = str(row.get("company", ""))
    product = str(row.get("product", ""))
    source = str(row.get("source", ""))
    if "동양생명" in company and all(num(row.get(key, 0)) == 0 for key in ("year1", "year2", "year3", "year4", "total")):
        return True
    summary_tokens = (
        "한국보험금융",
        "총계",
        "합계",
        "본부_",
        "채널_",
        "직영부문",
        "지점채널",
        "코인스",
    )
    if any(token in product for token in summary_tokens):
        return True
    if source == "직영파트 MS":
        return True
    return False


def clean(value: Any) -> str:
    return re.sub(r"\s+", " ", "" if value is None else str(value)).strip()


def compact(value: Any) -> str:
    return re.sub(r"\s+", "", clean(value))


def num(value: Any) -> float:
    if value is None or value == "":
        return 0.0
    if isinstance(value, (int, float)):
        return float(value)
    try:
        return float(str(value).replace(",", "").strip())
    except ValueError:
        return 0.0


def is_percent_format(number_format: str | None) -> bool:
    return "%" in str(number_format or "")


def format_percent(value: Any) -> str:
    amount = num(value) * 100
    if abs(amount) < 0.05:
        return "0%"
    text = f"{amount:.1f}".rstrip("0").rstrip(".")
    return f"{text}%"


def format_percent_point(value: Any) -> str:
    amount = num(value)
    if abs(amount) < 0.05:
        return "0%"
    text = f"{amount:.1f}".rstrip("0").rstrip(".")
    return f"{text}%"


def format_percent_point_fixed(value: Any) -> str:
    amount = num(value)
    if abs(amount) < 0.05:
        return "0%"
    return f"{amount:.1f}%"


def format_amount(value: Any) -> str:
    amount = num(value)
    if abs(amount) < 0.0000001:
        return "0"
    text = f"{amount:,.4f}".rstrip("0").rstrip(".")
    return text or "0"


def parse_percent_display(value: Any) -> float | None:
    display = clean(value)
    if not display.endswith("%"):
        return None
    try:
        return float(display[:-1].replace(",", "").strip())
    except ValueError:
        return None


def normalize_rate_displays(row: dict[str, Any]) -> dict[str, Any]:
    item = dict(row)
    item["insuranceType"] = item.get("insuranceType") or "life"
    item["metricMode"] = item.get("metricMode") or "percent"
    if item["insuranceType"] == "nonlife":
        item["metricMode"] = "percent"
        year_total = 0.0
        for key in ("year1", "year2", "year3", "year4"):
            value = num(item.get(key, 0))
            item[key] = value
            year_total += value
            item[f"{key}Display"] = format_percent_point_fixed(value)
        saved_total = num(item.get("total", 0))
        item["total"] = saved_total if saved_total else year_total
        item["totalDisplay"] = format_percent_point_fixed(item["total"])
        return item
    if item["metricMode"] == "amount":
        year_total = 0.0
        for key in ("year1", "year2", "year3", "year4"):
            value = num(item.get(key, 0))
            item[key] = value
            year_total += value
            item[f"{key}Display"] = format_amount(value)
        saved_total = num(item.get("total", 0))
        item["total"] = saved_total if saved_total else year_total
        item["totalDisplay"] = format_amount(item["total"])
        return item
    year_total = 0.0
    for key in ("year1", "year2", "year3", "year4"):
        display_value = parse_percent_display(item.get(f"{key}Display", ""))
        value = display_value if display_value is not None else num(item.get(key, 0))
        item[key] = value
        year_total += value
        item[f"{key}Display"] = format_percent_point(value)
    total_display_value = parse_percent_display(item.get("totalDisplay", ""))
    if total_display_value is not None:
        total = total_display_value
    else:
        saved_total = num(item.get("total", 0))
        total = saved_total if saved_total != 0 else year_total
    item["total"] = total
    item["totalDisplay"] = format_percent_point(total)
    return item


def row_is_header_noise(row: dict[str, Any]) -> bool:
    product = compact(row.get("product", ""))
    if not product:
        return True
    header_products = {
        "보종",
        "상품명",
        "상품구분",
        "구분",
        "납기",
        "납기(년)",
        "납입기간",
        "나이/보험료",
        "월납보험료",
        "보험료",
        "가입금액",
        "총환산월초(GAP)",
    }
    if product in header_products:
        return True
    if "보장금액" in product and "보험료" in product:
        return True
    return product.startswith("보종/") or product.startswith("상품명/")


CURRENT_UPLOAD: dict[str, Any] = load_saved_upload()


def cell_value(cell: Any) -> Any:
    return cell.value if hasattr(cell, "value") else cell


def cell_display(cell: Any) -> str:
    value = cell_value(cell)
    if value is None or value == "":
        return ""
    if isinstance(value, str):
        return clean(value)
    if hasattr(cell, "number_format") and is_percent_format(cell.number_format):
        return format_percent(value)
    return ""


def display_for(row: list[Any], col: int) -> str:
    if col < 0 or col >= len(row):
        return ""
    return cell_display(row[col])


def rate_value_for(row: list[Any], col: int) -> float:
    if col < 0 or col >= len(row):
        return 0.0
    cell = row[col]
    value = num(cell_value(cell))
    if hasattr(cell, "number_format") and is_percent_format(cell.number_format):
        return value * 100
    return value


def rate_display_for(row: list[Any], col: int) -> str:
    if col < 0 or col >= len(row):
        return "0%"
    return format_percent_point(rate_value_for(row, col))


def row_value(row: list[Any], col: int) -> Any:
    return cell_value(row[col]) if 0 <= col < len(row) else ""


def find_col(row: list[Any], candidates: list[str]) -> int:
    labels = [compact(cell_value(cell)) for cell in row]
    for idx, label in enumerate(labels):
        if any(candidate in label for candidate in candidates):
            return idx
    return -1


def find_year_col(rows: list[list[Any]], header_idx: int, candidates: list[str]) -> int:
    for offset in range(3):
        if header_idx + offset >= len(rows):
            continue
        for idx, cell in enumerate(rows[header_idx + offset]):
            label = compact(cell_value(cell))
            if any(candidate in label for candidate in candidates):
                return idx
    return -1


def parse_life_sheet(rows: list[list[Any]], sheet_name: str) -> list[dict[str, Any]]:
    header_idx = -1
    company_col = -1
    product_col = -1

    for idx, row in enumerate(rows[:25]):
        c_col = find_col(row, ["제휴사", "보험사", "회사"])
        p_col = find_col(row, ["상품구분", "상품명", "상품"])
        if c_col >= 0 and p_col >= 0:
            header_idx = idx
            company_col = c_col
            product_col = p_col
            break

    if header_idx < 0:
        return []

    y1_col = find_year_col(rows, header_idx, ["1차년", "1차년도"])
    y2_col = find_year_col(rows, header_idx, ["2차년", "2차년도"])
    y3_col = find_year_col(rows, header_idx, ["3차년", "3차년도"])
    y4_col = find_year_col(rows, header_idx, ["4차년", "4차년도"])
    if min(y1_col, y2_col, y3_col) < 0:
        return []

    parsed: list[dict[str, Any]] = []
    current_company = ""
    for row in rows[header_idx + 2 :]:
        company_cell = clean(cell_value(row[company_col]) if company_col < len(row) else "")
        product = clean(cell_value(row[product_col]) if product_col < len(row) else "")
        if company_cell:
            current_company = company_cell
        company = company_cell or current_company
        if not company or not product:
            continue
        if any(label in compact(product) for label in ["상품구분", "상품명"]):
            continue
        year1 = num(cell_value(row[y1_col]) if y1_col < len(row) else 0)
        year2 = num(cell_value(row[y2_col]) if y2_col < len(row) else 0)
        year3 = num(cell_value(row[y3_col]) if y3_col < len(row) else 0)
        year4 = num(cell_value(row[y4_col]) if y4_col >= 0 and y4_col < len(row) else 0)
        item = {
            "company": company,
            "product": product,
            "year1": year1,
            "year2": year2,
            "year3": year3,
            "year4": year4,
            "total": year1 + year2 + year3 + year4,
            "source": sheet_name,
        }
        for key, col in (("year1", y1_col), ("year2", y2_col), ("year3", y3_col), ("year4", y4_col)):
            display = display_for(row, col)
            if display:
                item[f"{key}Display"] = display
        parsed.append(item)
    return parsed


def company_from_sheet_name(sheet_name: str) -> str:
    return re.sub(r"\(.+\)$", "", sheet_name).strip() or sheet_name


def is_product_component_label(label: str) -> bool:
    return (
        label
        in {
            "보종",
            "상품명",
            "상품구분",
            "구분",
            "구분",
            "납입기간",
            "납기",
            "가입금액구간",
            "가입금액",
            "가입연령",
            "만기",
            "적용기준",
            "상품가입형태",
            "상품유형",
        }
        or "보장금액" in label
        or "보험료" in label
        or "보험기간" in label
        or "납입" in label
        or "납" in label
    )


def find_commission_rate_year_cols(rows: list[list[Any]], header_idx: int) -> tuple[int, int, int, int]:
    header = [compact(cell_value(cell)) for cell in rows[header_idx]]
    rate_col = -1
    for idx, label in enumerate(header):
        if "CommissionRate" in label or "수수료율" in label:
            rate_col = idx
            break
    if rate_col < 0:
        return -1, -1, -1, -1

    max_col = min(max(len(row) for row in rows[header_idx : min(len(rows), header_idx + 8)]), rate_col + 8)
    year_cols = [-1, -1, -1, -1]
    exact_labels = [
        {"1차년", "1차년도"},
        {"2차년", "2차년도"},
        {"3차년", "3차년도"},
        {"4차년", "4차년도", "4차년이후"},
    ]

    for row in rows[header_idx : min(len(rows), header_idx + 8)]:
        labels = [compact(cell_value(cell)) for cell in row]
        found = [-1, -1, -1, -1]
        for col in range(rate_col, max_col):
            label = labels[col] if col < len(labels) else ""
            for year_idx, candidates in enumerate(exact_labels):
                if label in candidates:
                    found[year_idx] = col
        if found[0] >= 0 and found[1] >= 0 and found[2] >= 0:
            year_cols = found
            break

    return tuple(year_cols)  # type: ignore[return-value]


def rate_percent(cell: Any) -> float:
    value = cell_value(cell)
    rate = num(value)
    if rate and abs(rate) < 1 and hasattr(cell, "number_format") and is_percent_format(cell.number_format):
        return rate * 100
    return rate


def sheet_component_cols(sheet_name: str, component_cols: list[int], first_year_col: int) -> list[int]:
    name = company_from_sheet_name(sheet_name)
    if "\uc0bc\uc131\uc0dd\uba85" in name:
        return [0, 1, 2]
    if "\ub3d9\uc591\uc0dd\uba85" in name:
        if first_year_col <= 5:
            return [0, 1, 2, 4]
        return [0, 1, 2, 3, 4, 5]
    overrides: dict[str, list[int]] = {
        "메트라이프": [0, 1, 2, 3],
        "처브라이프": [0, 1, 2, 3],
        "흥국생명": [0, 1, 2],
        "카디프생명": [0, 1, 2, 3],
    }
    for keyword, cols in overrides.items():
        if keyword in name:
            return [col for col in cols if col < first_year_col]
    return component_cols


def detail_sheet_override(sheet_name: str) -> dict[str, Any]:
    name = company_from_sheet_name(sheet_name)
    if "\uc0bc\uc131\uc0dd\uba85" in name:
        return {"years": [15, 23, 29, -1], "total": 31, "components": [0, 1, 2]}
    if "\ud55c\ud654\uc0dd\uba85" in name:
        return {"years": [14, (15, 22), (23, 29), (30, 38)], "total": 39, "components": [0, 2, 3]}
    if "\ub3d9\uc591\uc0dd\uba85" in name:
        return {"years": [20, 26, 32, -1], "total": 33, "components": [0, 1, 2, 3, 4, 5]}
    if "\uad50\ubcf4\uc0dd\uba85" in name:
        return {"years": [13, 20, 26, -1], "total": 27, "components": [0, 1, 2, 3]}
    if "\ud478\ubcf8\ud604\ub300" in name:
        return {"years": [9, 12, 14, -1], "total": 15, "components": [0, 1, 2, 3]}
    if "ABL\uc0dd\uba85" in name:
        return {"years": [4, (5, 6), 7, 8], "total": 9, "components": [0, 1, 2, 3]}
    if "IM\ub77c\uc774\ud504" in name or "iM\ub77c\uc774\ud504" in name:
        return {"years": [5, 12, 13, 14], "total": 15, "components": [0, 1, 2, 3]}
    return {}


def early_detail_sheet_override(sheet_name: str, year_cols: tuple[int, int, int, int], total_col: int) -> dict[str, Any]:
    name = company_from_sheet_name(sheet_name)
    y1_col, y2_col, y3_col, y4_col = year_cols
    if min(y1_col, y2_col, y3_col) < 0:
        return {}
    if y1_col > 12:
        return {}
    if "\uc0bc\uc131\uc0dd\uba85" in name and total_col == 8:
        return {"years": [y1_col, y2_col, y3_col, y4_col], "total": total_col, "components": [0, 1, 2]}
    if "\ud55c\ud654\uc0dd\uba85" in name and total_col == 13:
        return {"years": [y1_col, y2_col, y3_col, y4_col], "total": total_col, "components": [0, 2, 3]}
    if "\ub3d9\uc591\uc0dd\uba85" in name and total_col == 8:
        return {"years": [y1_col, y2_col, y3_col, y4_col], "total": total_col, "components": [0, 1, 2, 4]}
    if "\ub3d9\uc591\uc0dd\uba85" in name and total_col == 15:
        return {"years": [y1_col, y2_col, y3_col, y4_col], "total": total_col, "components": [0, 1, 2, 3, 4, 5]}
    if "\uad50\ubcf4\uc0dd\uba85" in name and total_col == 7:
        return {"years": [y1_col, y2_col, y3_col, y4_col], "total": total_col, "components": [0, 1, 2, 3]}
    if "\ud478\ubcf8\ud604\ub300" in name and total_col == 7:
        return {"years": [y1_col, y2_col, y3_col, y4_col], "total": total_col, "components": [0, 1, 2, 3]}
    return {}


def rate_value_from_spec(row: list[Any], spec: Any) -> float:
    if isinstance(spec, tuple):
        start, end = spec
        return sum(rate_value_for(row, col) for col in range(start, end + 1))
    if isinstance(spec, int):
        return rate_value_for(row, spec)
    return 0.0


def rate_display_from_spec(row: list[Any], spec: Any) -> str:
    return format_percent_point(rate_value_from_spec(row, spec))


def component_text(cell: Any) -> str:
    value = cell_value(cell)
    text = clean(value)
    if value is None or value == "":
        return ""
    if isinstance(value, (int, float)):
        number_text = f"{float(value):.10g}"
        number_format = str(getattr(cell, "number_format", "") or "")
        literals = "".join(re.findall(r'"([^"]+)"', number_format))
        if literals:
            return f"{number_text}{literals}"
        return number_text
    return text


def build_product(sheet_name: str, parts: list[str]) -> str:
    deduped: list[str] = []
    for part in parts:
        if part and (not deduped or compact(deduped[-1]) != compact(part)):
            deduped.append(part)
    parts = deduped
    if not parts:
        return ""
    company = company_from_sheet_name(sheet_name)
    if "흥국생명" in company and len(parts) >= 3:
        return f"{parts[0]}({parts[1]} {parts[2]})"
    return " / ".join(parts)


def is_header_component_value(value: str) -> bool:
    label = compact(value)
    if not label:
        return True
    if label in {"-", "\u2013", "\u2014"}:
        return True
    return label in {
        "보종",
        "상품명",
        "상품구분",
        "구분",
        "납입기간",
        "납기",
        "납기(년)",
        "나이/보험료",
        "가입금액구간",
        "가입금액",
        "가입연령",
        "만기",
        "적용기준",
        "상품가입형태",
        "상품유형",
        "보험기간",
        "주피연령",
        "환산율",
    }


def find_detail_year_cols(rows: list[list[Any]], header_idx: int) -> tuple[int, int, int, int]:
    search_rows = rows[header_idx : min(len(rows), header_idx + 6)]

    def direct_year_cols(year: int) -> list[int]:
        matches: list[int] = []
        for row in search_rows:
            for col, cell in enumerate(row):
                label = compact(cell_value(cell))
                if label in {f"{year}차년", f"{year}차년도"}:
                    matches.append(col)
        return matches

    def matching_cols(year: int, require_total: bool) -> list[int]:
        matches: list[int] = []
        for row in search_rows:
            for col, cell in enumerate(row):
                label = compact(cell_value(cell))
                if not label:
                    continue
                has_year = bool(re.search(rf"{year}(?:~\d+)?차년(?:도)?", label))
                has_first_year = year == 1 and ("초년도" in label or "초년" in label)
                has_first_year_summary = year == 1 and any(
                    token in label for token in ["익월계", "익월計", "익월총", "초년도계", "초년도計"]
                )
                if not has_year and not has_first_year and not has_first_year_summary:
                    continue
                if "초과" in label:
                    continue
                has_total = "계" in label or "計" in label or has_first_year_summary
                if require_total and not has_total:
                    continue
                matches.append(col)
        return matches

    cols: list[int] = []
    for year in [1, 2, 3, 4]:
        direct = direct_year_cols(year)
        preferred = matching_cols(year, True)
        fallback = matching_cols(year, False)
        cols.append(direct[0] if direct else (preferred or fallback or [-1])[-1])
    return tuple(cols)  # type: ignore[return-value]


def find_detail_total_col(rows: list[list[Any]], header_idx: int) -> int:
    for row in rows[header_idx : min(len(rows), header_idx + 6)]:
        for col, cell in enumerate(row):
            label = compact(cell_value(cell))
            if "총수수료" in label:
                return col
    return -1


def parse_life_company_detail_sheet(rows: list[list[Any]], sheet_name: str) -> list[dict[str, Any]]:
    header_idx = -1
    year_cols = (-1, -1, -1, -1)
    override = detail_sheet_override(sheet_name)

    for idx, _row in enumerate(rows[:35]):
        found_year_cols = find_detail_year_cols(rows, idx)
        if min(found_year_cols[:3]) < 0:
            continue
        if not (found_year_cols[0] <= found_year_cols[1] <= found_year_cols[2]):
            continue
        header_rows = rows[idx : min(len(rows), idx + 6)]
        first_year_col = min(col for col in found_year_cols[:3] if col >= 0)
        if first_year_col <= 0:
            continue
        has_product_label = any(
            is_product_component_label(compact(cell_value(cell)))
            for row in header_rows
            for cell in row[:first_year_col]
        )
        if has_product_label:
            header_idx = idx
            year_cols = found_year_cols
            break

    if header_idx < 0:
        return []

    detected_total_col = find_detail_total_col(rows, header_idx)
    override = early_detail_sheet_override(sheet_name, year_cols, detected_total_col) or override
    first_year_col = min(col for col in year_cols[:3] if col >= 0)
    header_rows = rows[header_idx : min(len(rows), header_idx + 6)]
    component_cols: list[int] = []
    for col in range(first_year_col):
        labels = [compact(cell_value(row[col]) if col < len(row) else "") for row in header_rows]
        if any(is_product_component_label(label) for label in labels):
            component_cols.append(col)

    if not component_cols:
        component_cols = list(range(min(first_year_col, 4)))
    if override.get("components"):
        component_cols = [col for col in override["components"] if col < first_year_col]
    component_cols = sheet_component_cols(sheet_name, component_cols, first_year_col)
    year_specs = override.get("years", list(year_cols))
    total_col = override.get("total", detected_total_col)

    company = company_from_sheet_name(sheet_name)
    parsed: list[dict[str, Any]] = []
    y1_col, y2_col, y3_col, y4_col = year_cols

    for row in rows[header_idx + len(header_rows) :]:
        row_parts: dict[int, str] = {}
        for col in component_cols:
            cell = component_text(row[col]) if col < len(row) else ""
            compact_cell = compact(cell)
            if not cell:
                continue
            if is_header_component_value(compact_cell) or compact_cell == "환산율":
                continue
            row_parts[col] = cell

        product = build_product(sheet_name, [row_parts.get(col, "") for col in component_cols])
        if not product:
            continue
        if compact(product) in {"월납보험료", "보험료", "가입금액", "총환산월초(GAP)"}:
            continue
        if all(is_header_component_value(part) for part in product.split(" / ")):
            continue

        year1 = rate_value_from_spec(row, year_specs[0])
        year2 = rate_value_from_spec(row, year_specs[1])
        year3 = rate_value_from_spec(row, year_specs[2])
        year4 = rate_value_from_spec(row, year_specs[3])
        total_from_sheet = rate_value_for(row, total_col)
        has_any_rate_value = any(
            rate_value_for(row, col) != 0
            for col in range(first_year_col, min(len(row), first_year_col + 16))
        )
        if year1 == 0 and year2 == 0 and year3 == 0 and year4 == 0 and not has_any_rate_value:
            continue

        total = total_from_sheet if total_col >= 0 and total_from_sheet != 0 else year1 + year2 + year3 + year4
        item = {
            "company": company,
            "product": product,
            "year1": year1,
            "year2": year2,
            "year3": year3,
            "year4": year4,
            "total": total,
            "source": sheet_name,
        }
        for key, spec in (("year1", year_specs[0]), ("year2", year_specs[1]), ("year3", year_specs[2]), ("year4", year_specs[3])):
            item[f"{key}Display"] = rate_display_from_spec(row, spec)
        item["totalDisplay"] = rate_display_for(row, total_col) if total_col >= 0 and total_from_sheet != 0 else format_percent_point(item["total"])
        parsed.append(item)

    return parsed


def parse_life_commission_rate_sheet(rows: list[list[Any]], sheet_name: str) -> list[dict[str, Any]]:
    header_idx = -1
    component_cols: list[tuple[int, str]] = []

    for idx, row in enumerate(rows[:30]):
        labels = [compact(cell_value(cell)) for cell in row]
        if "상품명" in labels and any("CommissionRate" in label or "수수료율" in label for label in labels):
            header_idx = idx
            component_cols = [
                (col_idx, label)
                for col_idx, label in enumerate(labels)
                if is_product_component_label(label)
            ]
            break

    if header_idx < 0 or not component_cols:
        return []

    override_cols = sheet_component_cols(sheet_name, [col_idx for col_idx, _label in component_cols], len(rows[header_idx]))
    if override_cols:
        label_by_col = {col_idx: label for col_idx, label in component_cols}
        component_cols = [(col_idx, label_by_col.get(col_idx, "")) for col_idx in override_cols]

    y1_col, y2_col, y3_col, y4_col = find_commission_rate_year_cols(rows, header_idx)
    if min(y1_col, y2_col, y3_col) < 0:
        return []
    total_col = find_detail_total_col(rows, header_idx)
    first_year_col = min(col for col in (y1_col, y2_col, y3_col) if col >= 0)

    company = company_from_sheet_name(sheet_name)
    current_parts: dict[int, str] = {}
    parsed: list[dict[str, Any]] = []

    for row in rows[header_idx + 1 :]:
        for col_idx, _label in component_cols:
            cell = clean(cell_value(row[col_idx]) if col_idx < len(row) else "")
            if cell and not is_header_component_value(cell):
                current_parts[col_idx] = cell

        parts = [current_parts.get(col_idx, "") for col_idx, _label in component_cols]
        product = build_product(sheet_name, parts)
        if not product or "상품명" in compact(product) or "CommissionRate" in compact(product):
            continue
        if all(is_header_component_value(part) for part in product.split(" / ")):
            continue

        year1 = rate_value_for(row, y1_col)
        year2 = rate_value_for(row, y2_col)
        year3 = rate_value_for(row, y3_col)
        year4 = rate_value_for(row, y4_col)
        total_from_sheet = rate_value_for(row, total_col)
        has_any_rate_value = any(
            rate_value_for(row, col) != 0
            for col in range(first_year_col, min(len(row), first_year_col + 16))
        )
        if year1 == 0 and year2 == 0 and year3 == 0 and year4 == 0 and not has_any_rate_value:
            continue

        total = total_from_sheet if total_col >= 0 and total_from_sheet != 0 else year1 + year2 + year3 + year4
        item = {
            "company": company,
            "product": product,
            "year1": year1,
            "year2": year2,
            "year3": year3,
            "year4": year4,
            "total": total,
            "source": sheet_name,
        }
        for key, col in (("year1", y1_col), ("year2", y2_col), ("year3", y3_col), ("year4", y4_col)):
            item[f"{key}Display"] = rate_display_for(row, col)
        item["totalDisplay"] = rate_display_for(row, total_col) if total_col >= 0 and total_from_sheet != 0 else format_percent_point(item["total"])
        parsed.append(item)

    return parsed


def parse_nonlife_sheet(rows: list[list[Any]], sheet_name: str) -> list[dict[str, Any]]:
    header_idx = -1
    for idx, row in enumerate(rows):
        labels = [compact(cell_value(cell)) for cell in row]
        if "보험사" in labels and "상품명" in labels and any("총환산" in label for label in labels):
            header_idx = idx
            break
    if header_idx < 0:
        return []

    header = rows[header_idx]
    company_col = find_col(header, ["보험사"])
    product_col = find_col(header, ["상품명"])
    y1_col = find_col(header, ["총환산"])
    y2_col = find_col(header, ["환산2차년도", "2차년도", "2차년"])
    y3_col = find_col(header, ["환산3차년도", "3차년도", "3차년"])
    y4_col = find_col(header, ["환산4차년도", "4차년도", "4차년"])
    if min(company_col, product_col, y1_col, y2_col, y3_col) < 0:
        return []

    grouped: dict[tuple[str, str], dict[str, Any]] = {}
    for row in rows[header_idx + 1 :]:
        company = clean(cell_value(row[company_col]) if company_col < len(row) else "")
        product = clean(cell_value(row[product_col]) if product_col < len(row) else "")
        if not company or not product:
            continue
        key = (company, product)
        item = grouped.setdefault(
            key,
            {"company": company, "product": product, "year1": 0.0, "year2": 0.0, "year3": 0.0, "year4": 0.0, "source": f"{sheet_name} 합산"},
        )
        for key, col in (("year1", y1_col), ("year2", y2_col), ("year3", y3_col), ("year4", y4_col)):
            cell = row[col] if col >= 0 and col < len(row) else 0
            item[key] += num(cell_value(cell))
            if hasattr(cell, "number_format") and is_percent_format(cell.number_format):
                item[f"_{key}IsPercent"] = True

    result = []
    for item in grouped.values():
        item["total"] = item["year1"] + item["year2"] + item["year3"] + item.get("year4", 0.0)
        for key in ("year1", "year2", "year3", "year4"):
            if item.pop(f"_{key}IsPercent", False):
                item[f"{key}Display"] = format_percent(item[key])
        result.append(item)
    return result


def parse_nonlife_amount_sheet(rows: list[list[Any]], sheet_name: str) -> list[dict[str, Any]]:
    header_idx = -1
    for idx, row in enumerate(rows[:80]):
        labels = [compact(cell_value(cell)) for cell in row]
        has_product = any("상품" in label or "담보" in label for label in labels)
        has_next_month = any("익월" in label for label in labels)
        has_second_year = any("2차" in label or "2차년도" in label or "13회" in label for label in labels)
        has_total = any("총수수료" in label or "총수수료액" in label or "총수수료금액" in label for label in labels)
        if has_product and has_next_month and (has_second_year or has_total):
            header_idx = idx
            break
    if header_idx < 0:
        return []

    header = rows[header_idx]
    labels = [compact(cell_value(cell)) for cell in header]

    def find_label(candidates: list[str], fallback: int = -1) -> int:
        for col, label in enumerate(labels):
            if any(candidate in label for candidate in candidates):
                return col
        return fallback

    company_col = find_label(["보험사", "회사", "제휴사"])
    product_col = find_label(["상품명", "상품", "담보명", "담보"])
    next_month_col = find_label(["익월"])
    second_year_col = -1
    for col, label in enumerate(labels):
        if "2차" in label or "2차년도" in label or "2차년" in label or label == "13회":
            second_year_col = col
            break
    total_col = find_label(["총수수료", "총수수료액", "총수수료금액", "합계"])
    if product_col < 0 or next_month_col < 0 or second_year_col < 0:
        return []

    sub_labels = [compact(cell_value(cell)) for cell in rows[header_idx + 1]] if header_idx + 1 < len(rows) else []
    detail_cols: list[int] = []
    for col in range(product_col + 1, next_month_col):
        label = labels[col] if col < len(labels) else ""
        sub_label = sub_labels[col] if col < len(sub_labels) else ""
        merged_label = label or sub_label
        if not merged_label or any(token in merged_label for token in ["수정율", "수정률", "환산율", "수수료", "익월"]):
            continue
        detail_cols.append(col)
    second_year_cols = [
        col
        for col in range(second_year_col, total_col if total_col > second_year_col else len(header))
        if col != next_month_col
    ] or [second_year_col]

    parsed: list[dict[str, Any]] = []
    current_company = sheet_company(sheet_name)
    current_product = ""
    current_details: dict[int, str] = {}
    for row in rows[header_idx + 1 :]:
        company = clean(row_value(row, company_col)) if 0 <= company_col < len(row) else ""
        product = clean(row_value(row, product_col)) if product_col < len(row) else ""
        if company and compact(company) not in {"보험사", "회사", "제휴사"}:
            current_company = company
        if product and compact(product) not in {"상품", "상품명", "담보", "담보명"}:
            current_product = product
        company = current_company
        product = product or current_product
        details: list[str] = []
        for col in detail_cols:
            detail = clean(row_value(row, col)) if col < len(row) else ""
            if detail and compact(detail) not in {"구분", "납입기간", "보험기간", "담보명", "담보"}:
                current_details[col] = detail
            else:
                detail = current_details.get(col, "")
            if detail:
                details.append(detail)
        product = build_product(sheet_name, [product, *details])
        if not company or not product:
            continue
        if any(token in compact(product) for token in ["합계", "총계"]) and len(compact(product)) <= 8:
            continue
        next_month = num(row_value(row, next_month_col)) * 100
        second_year = sum(num(row_value(row, col)) for col in second_year_cols if col < len(row)) * 100
        total = (num(row_value(row, total_col)) * 100) if total_col >= 0 else next_month + second_year
        if next_month == 0 and second_year == 0 and total == 0:
            continue
        item = {
            "company": company,
            "product": product,
            "year1": next_month,
            "year2": second_year,
            "year3": 0.0,
            "year4": 0.0,
            "total": total if total else next_month + second_year,
            "source": sheet_name,
            "insuranceType": "nonlife",
            "metricMode": "percent",
        }
        item["year1Display"] = format_percent_point_fixed(item["year1"])
        item["year2Display"] = format_percent_point_fixed(item["year2"])
        item["year3Display"] = "0%"
        item["year4Display"] = "0%"
        item["totalDisplay"] = format_percent_point_fixed(item["total"])
        parsed.append(item)
    return parsed


def parse_monthly_matrix_sheet(rows: list[list[Any]], sheet_name: str) -> list[dict[str, Any]]:
    month_header_idx = -1
    month_cols: list[tuple[int, str]] = []

    for idx, row in enumerate(rows[:20]):
        labels = [compact(cell_value(cell)) for cell in row]
        candidate_cols = [col for col, label in enumerate(labels) if label == "월납"]
        if not candidate_cols:
            continue
        company_row = rows[idx - 1] if idx > 0 else []
        for col in candidate_cols:
            company = ""
            for offset in range(col, max(-1, col - 3), -1):
                if offset < len(company_row):
                    company = clean(cell_value(company_row[offset]))
                    if company and company not in {"값", "보험사"}:
                        break
            if company and company not in {"값", "보험사"}:
                month_cols.append((col, company))
        if month_cols:
            month_header_idx = idx
            break

    if month_header_idx < 0:
        return []

    first_month_col = min(col for col, _company in month_cols)
    dimension_names = {"채널", "채널명", "본부", "본부명", "사업부", "사업부명", "사업단명", "지사명", "지점명"}
    dimension_cols = [
        col
        for col, cell in enumerate(rows[month_header_idx][:first_month_col])
        if compact(cell_value(cell)) in dimension_names
    ]
    if not dimension_cols and month_header_idx > 0:
        dimension_cols = [
            col
            for col, cell in enumerate(rows[month_header_idx - 1][:first_month_col])
            if compact(cell_value(cell)) in dimension_names
        ]
    if not dimension_cols:
        dimension_cols = list(range(min(3, first_month_col)))

    parsed: list[dict[str, Any]] = []
    for row in rows[month_header_idx + 1 :]:
        org_parts = [clean(cell_value(row[col]) if col < len(row) else "") for col in dimension_cols]
        product = " / ".join(part for part in org_parts if part)
        if not product:
            continue
        if any(skip in product for skip in ["총계", "합계"]) and len(product) <= 12:
            product = product
        for col, company in month_cols:
            amount = num(cell_value(row[col]) if col < len(row) else 0)
            if amount == 0:
                continue
            item = {
                "company": company,
                "product": product,
                "year1": amount,
                "year2": 0.0,
                "year3": 0.0,
                "year4": 0.0,
                "total": amount,
                "source": sheet_name,
            }
            display = display_for(row, col)
            if display:
                item["year1Display"] = display
            parsed.append(item)
    return parsed


def sheet_company(sheet_name: str) -> str:
    name = re.sub(r"\([^)]*\)", "", sheet_name).strip()
    return name or sheet_name


def header_text(rows: list[list[Any]], header_idx: int, col: int) -> str:
    parts = []
    for idx in range(max(0, header_idx - 3), min(len(rows), header_idx + 3)):
        value = clean(row_value(rows[idx], col))
        if value:
            parts.append(value)
    return compact(" ".join(parts))


def has_percent_number(row: list[Any], col: int) -> bool:
    if col < 0 or col >= len(row):
        return False
    cell = row[col]
    value = cell_value(cell)
    return isinstance(value, (int, float)) and hasattr(cell, "number_format") and is_percent_format(cell.number_format)


def likely_header_score(row: list[Any]) -> int:
    labels = [compact(cell_value(cell)) for cell in row if clean(cell_value(cell))]
    score = 0
    for label in labels:
        if any(token in label for token in ["상품", "회사", "보험", "구분", "수수료", "환산", "1차", "2차", "3차", "4차"]):
            score += 2
        if any(token in label for token in ["惑前", "蹂댄뿕", "备盒", "荐荐丰", "券魂", "1瞒", "2瞒", "3瞒", "4瞒"]):
            score += 2
        if any(token.lower() in label.lower() for token in ["commission", "rate", "product"]):
            score += 2
    return score


def find_generic_header(rows: list[list[Any]]) -> int:
    best_idx = -1
    best_score = 0
    for idx, row in enumerate(rows[:30]):
        score = likely_header_score(row)
        percent_count = sum(1 for cell in row if hasattr(cell, "number_format") and is_percent_format(cell.number_format))
        if percent_count >= 3:
            score += percent_count
        if score > best_score:
            best_idx = idx
            best_score = score
    return best_idx if best_score >= 4 else -1


def find_product_col(rows: list[list[Any]], header_idx: int, first_rate_col: int) -> int:
    max_col = max(1, min(first_rate_col, 8))
    hard_disallowed = ["상품구분", "상품유형", "상품가입"]
    disallowed = hard_disallowed + ["구분", "유형", "납기", "납입", "기간", "만기", "연령", "보험료", "환산", "수수료", "성과"]

    for col in range(max_col):
        label = header_text(rows, header_idx, col)
        if "상품명" in label and not any(token in label for token in hard_disallowed):
            return col

    scores: dict[int, float] = {col: 0.0 for col in range(max_col)}
    for col in range(max_col):
        label = header_text(rows, header_idx, col)
        if any(token in label for token in ["상품명", "보험명"]):
            scores[col] += 60
        if any(token in label for token in ["보험", "무배당", "종신", "연금", "보장"]):
            scores[col] += 25
        if any(token in label for token in disallowed):
            scores[col] -= 80
        if label.startswith(("※", "∝", "*")):
            scores[col] -= 120

    for row in rows[header_idx + 1 : min(len(rows), header_idx + 140)]:
        if sum(1 for col in range(len(row)) if has_percent_number(row, col)) < 1:
            continue
        for col in range(max_col):
            text = clean(row_value(row, col))
            if len(text) < 2 or re.fullmatch(r"[-\d.,% ]+", text):
                continue
            compact_text = compact(text)
            score = min(len(text), 60)
            if any(token in compact_text for token in ["보험", "무배당", "종신", "연금", "보장", "라이프"]):
                score += 45
            if any(token in compact_text for token in ["상품구분", "상품유형", "구분", "유형", "납기", "납입", "기간", "만기", "연령", "보험료"]):
                score -= 70
            if len(text) <= 5:
                score -= 15
            scores[col] += score
    return max(scores, key=scores.get) if scores else 0


def find_rate_col(rows: list[list[Any]], header_idx: int, tokens: list[str], used: set[int]) -> int:
    candidates: list[tuple[int, int]] = []
    width = max((len(row) for row in rows[: min(len(rows), 40)]), default=0)
    for col in range(width):
        if col in used:
            continue
        label = header_text(rows, header_idx, col)
        normalized_tokens = tokens + [token.encode("utf-8").decode("utf-8") for token in tokens]
        if not any(token in label for token in normalized_tokens):
            continue
        count = sum(1 for row in rows[header_idx + 1 : min(len(rows), header_idx + 120)] if has_percent_number(row, col))
        if count:
            candidates.append((count, col))
    if not candidates:
        return -1
    candidates.sort(key=lambda item: (-item[0], item[1]))
    return candidates[0][1]


def parse_generic_rate_sheet(rows: list[list[Any]], sheet_name: str) -> list[dict[str, Any]]:
    header_idx = find_generic_header(rows)
    if header_idx < 0:
        return []

    percent_cols = sorted(
        {
            col
            for row in rows[header_idx + 1 : min(len(rows), header_idx + 120)]
            for col in range(len(row))
            if has_percent_number(row, col)
        }
    )
    if not percent_cols:
        return []

    first_rate_col = min(percent_cols)
    product_col = find_product_col(rows, header_idx, first_rate_col)
    company = sheet_company(sheet_name)

    used: set[int] = set()
    y1_col = find_rate_col(rows, header_idx, ["\u0031\ucc28", "1瞒", "\ucd08\ub144", "\ucd08\ud68c", "CommissionRate"], used)
    if y1_col >= 0:
        used.add(y1_col)
    y2_col = find_rate_col(rows, header_idx, ["\u0032\ucc28", "2瞒"], used)
    if y2_col >= 0:
        used.add(y2_col)
    y3_col = find_rate_col(rows, header_idx, ["\u0033\ucc28", "3瞒"], used)
    if y3_col >= 0:
        used.add(y3_col)
    y4_col = find_rate_col(rows, header_idx, ["\u0034\ucc28", "4瞒"], used)
    if y4_col >= 0:
        used.add(y4_col)

    selected = [col for col in (y1_col, y2_col, y3_col, y4_col) if col >= 0]
    if not selected:
        selected = [col for col in percent_cols if col >= product_col][:4]
        y1_col, y2_col, y3_col, y4_col = (selected + [-1, -1, -1, -1])[:4]
    elif y1_col < 0:
        y1_col = selected[0]

    parsed: list[dict[str, Any]] = []
    current_product = ""
    for row in rows[header_idx + 1 :]:
        raw_product = clean(row_value(row, product_col))
        base_product = raw_product or current_product
        if not base_product or len(base_product) < 2:
            continue
        compact_product = compact(base_product)
        if compact_product in {"상품", "상품명", "회사", "보험사", "수수료", "구분", "惑前疙", "荐荐丰", "券魂"}:
            continue
        if raw_product:
            current_product = raw_product
        if not any(has_percent_number(row, col) for col in (y1_col, y2_col, y3_col, y4_col)):
            continue

        category_parts = []
        dimension_limit = min([col for col in (y1_col, y2_col, y3_col, y4_col, first_rate_col) if col >= 0] or [first_rate_col])
        for col in range(0, max(dimension_limit, product_col + 1)):
            if col == product_col:
                continue
            text = clean(row_value(row, col))
            if not text or text == base_product:
                continue
            if text in {"-", "–", "—"}:
                continue
            category_parts.append(text[:80])
        category = " / ".join(category_parts[:5])
        product = base_product if not category else f"{base_product} ({category})"

        year1 = num(row_value(row, y1_col))
        year2 = num(row_value(row, y2_col))
        year3 = num(row_value(row, y3_col))
        year4 = num(row_value(row, y4_col))
        item = {
            "company": company,
            "product": product,
            "category": category,
            "year1": year1,
            "year2": year2,
            "year3": year3,
            "year4": year4,
            "total": year1 + year2 + year3 + year4,
            "source": sheet_name,
        }
        for key, col in (("year1", y1_col), ("year2", y2_col), ("year3", y3_col), ("year4", y4_col)):
            display = display_for(row, col)
            if display:
                item[f"{key}Display"] = display
        parsed.append(item)
    return parsed


def worksheet_rows(ws: Any) -> list[list[Any]]:
    max_col = min(ws.max_column or 1, MAX_PARSE_COLS)
    merged_lookup: dict[tuple[int, int], Any] = {}
    for merged_range in getattr(ws, "merged_cells", ()).ranges:
        if merged_range.min_col > max_col:
            continue
        top_left = ws.cell(merged_range.min_row, merged_range.min_col)
        for row_idx in range(merged_range.min_row, merged_range.max_row + 1):
            for col_idx in range(merged_range.min_col, min(merged_range.max_col, max_col) + 1):
                if row_idx == merged_range.min_row and col_idx == merged_range.min_col:
                    continue
                merged_lookup[(row_idx, col_idx)] = top_left

    rows: list[list[Any]] = []
    for row in ws.iter_rows(max_col=max_col):
        rows.append([merged_lookup.get((cell.row, cell.column), cell) for cell in row])
    return rows


FREE_SECT = 0xFFFFFFFF
END_OF_CHAIN = 0xFFFFFFFE


def _u16(data: bytes, offset: int) -> int:
    return struct.unpack_from("<H", data, offset)[0]


def _u32(data: bytes, offset: int) -> int:
    return struct.unpack_from("<I", data, offset)[0]


def _sector(data: bytes, sector_size: int, sid: int) -> bytes:
    start = 512 + sid * sector_size
    return data[start : start + sector_size]


def _chain(fat: list[int], start: int) -> list[int]:
    result: list[int] = []
    sid = start
    seen: set[int] = set()
    while sid not in (FREE_SECT, END_OF_CHAIN) and 0 <= sid < len(fat) and sid not in seen:
        seen.add(sid)
        result.append(sid)
        sid = fat[sid]
    return result


def _read_stream(data: bytes, sector_size: int, fat: list[int], start: int, size: int) -> bytes:
    payload = b"".join(_sector(data, sector_size, sid) for sid in _chain(fat, start))
    return payload[:size]


def _parse_cfb_streams(data: bytes) -> dict[str, bytes]:
    sector_size = 1 << _u16(data, 30)
    mini_sector_size = 1 << _u16(data, 32)
    num_fat = _u32(data, 44)
    first_dir = _u32(data, 48)
    mini_cutoff = _u32(data, 56)
    first_mini_fat = _u32(data, 60)
    num_mini_fat = _u32(data, 64)
    first_difat = _u32(data, 68)
    num_difat = _u32(data, 72)

    difat = [_u32(data, 76 + i * 4) for i in range(109)]
    sid = first_difat
    for _ in range(num_difat):
        block = _sector(data, sector_size, sid)
        entries = sector_size // 4 - 1
        difat.extend(_u32(block, i * 4) for i in range(entries))
        sid = _u32(block, entries * 4)
    fat_sids = [item for item in difat if item not in (FREE_SECT, END_OF_CHAIN)][:num_fat]
    fat: list[int] = []
    for sid in fat_sids:
        block = _sector(data, sector_size, sid)
        fat.extend(_u32(block, i * 4) for i in range(sector_size // 4))

    dir_stream = _read_stream(data, sector_size, fat, first_dir, 10**9)
    entries: list[tuple[str, int, int, int]] = []
    for offset in range(0, len(dir_stream), 128):
        entry = dir_stream[offset : offset + 128]
        if len(entry) < 128:
            break
        name_len = _u16(entry, 64)
        if name_len < 2:
            continue
        name = entry[: name_len - 2].decode("utf-16le", errors="ignore")
        entry_type = entry[66]
        start = _u32(entry, 116)
        size = struct.unpack_from("<Q", entry, 120)[0]
        entries.append((name, entry_type, start, size))

    root = next((entry for entry in entries if entry[1] == 5), None)
    mini_stream = _read_stream(data, sector_size, fat, root[2], root[3]) if root else b""
    mini_fat: list[int] = []
    if first_mini_fat not in (FREE_SECT, END_OF_CHAIN):
        for sid in _chain(fat, first_mini_fat)[:num_mini_fat]:
            block = _sector(data, sector_size, sid)
            mini_fat.extend(_u32(block, i * 4) for i in range(sector_size // 4))

    def read_mini(start: int, size: int) -> bytes:
        chunks: list[bytes] = []
        sid = start
        seen: set[int] = set()
        while sid not in (FREE_SECT, END_OF_CHAIN) and 0 <= sid < len(mini_fat) and sid not in seen:
            seen.add(sid)
            begin = sid * mini_sector_size
            chunks.append(mini_stream[begin : begin + mini_sector_size])
            sid = mini_fat[sid]
        return b"".join(chunks)[:size]

    streams: dict[str, bytes] = {}
    for name, entry_type, start, size in entries:
        if entry_type != 2:
            continue
        if size < mini_cutoff and mini_stream and mini_fat:
            streams[name] = read_mini(start, size)
        else:
            streams[name] = _read_stream(data, sector_size, fat, start, size)
    return streams


def _biff_records(stream: bytes, start: int = 0):
    pos = start
    while pos + 4 <= len(stream):
        code, length = struct.unpack_from("<HH", stream, pos)
        pos += 4
        payload = stream[pos : pos + length]
        pos += length
        yield code, payload, pos


def _decode_short_biff_string(payload: bytes, offset: int, encoding: str) -> str:
    length = payload[offset]
    flags = payload[offset + 1]
    offset += 2
    is_wide = bool(flags & 1)
    raw_len = length * (2 if is_wide else 1)
    raw = payload[offset : offset + raw_len]
    return raw.decode("utf-16le" if is_wide else encoding, errors="ignore")


def _decode_sst_string(buf: bytes, offset: int, encoding: str) -> tuple[str, int]:
    length = _u16(buf, offset)
    flags = buf[offset + 2]
    offset += 3
    rich_count = _u16(buf, offset) if flags & 0x08 else 0
    if flags & 0x08:
        offset += 2
    ext_len = _u32(buf, offset) if flags & 0x04 else 0
    if flags & 0x04:
        offset += 4
    is_wide = bool(flags & 0x01)
    raw_len = length * (2 if is_wide else 1)
    raw = buf[offset : offset + raw_len]
    offset += raw_len + rich_count * 4 + ext_len
    return raw.decode("utf-16le" if is_wide else encoding, errors="ignore"), offset


def _rk_value(raw: int) -> float:
    divide_100 = bool(raw & 1)
    is_int = bool(raw & 2)
    value_raw = raw & 0xFFFFFFFC
    if is_int:
        if value_raw & 0x80000000:
            value_raw -= 0x100000000
        value: float = value_raw >> 2
    else:
        value = struct.unpack("<d", struct.pack("<Q", value_raw << 32))[0]
    return value / 100 if divide_100 else value


def parse_xls_rows_without_xlrd(file_bytes: bytes) -> list[tuple[str, list[list[Any]]]]:
    streams = _parse_cfb_streams(file_bytes)
    workbook_stream = streams.get("Workbook") or streams.get("Book")
    if not workbook_stream:
        raise ValueError(".xls 통합 문서 데이터를 찾지 못했습니다.")

    bounds: list[tuple[str, int]] = []
    sst: list[str] = []
    encoding = "cp949"
    records = list(_biff_records(workbook_stream))
    for idx, (code, payload, _pos) in enumerate(records):
        if code == 0x0042 and len(payload) >= 2:
            codepage = _u16(payload, 0)
            encoding = "cp949" if codepage in (949, 0x8000) else "latin1"
        elif code == 0x0085 and len(payload) >= 8:
            bounds.append((_decode_short_biff_string(payload, 6, encoding), _u32(payload, 0)))
        elif code == 0x00FC and len(payload) >= 8:
            chunks = [payload]
            next_idx = idx + 1
            while next_idx < len(records) and records[next_idx][0] == 0x003C:
                chunks.append(records[next_idx][1])
                next_idx += 1
            buf = b"".join(chunks)
            total = _u32(buf, 4)
            offset = 8
            for _ in range(total):
                if offset >= len(buf):
                    break
                try:
                    text, offset = _decode_sst_string(buf, offset, encoding)
                except Exception:
                    break
                sst.append(text)

    result: list[tuple[str, list[list[Any]]]] = []
    for sheet_index, (sheet_name, start) in enumerate(bounds):
        end = bounds[sheet_index + 1][1] if sheet_index + 1 < len(bounds) else len(workbook_stream)
        cells: dict[int, dict[int, Any]] = {}
        merges: list[tuple[int, int, int, int]] = []
        for code, payload, _pos in _biff_records(workbook_stream[start:end]):
            if code == 0x00FD and len(payload) >= 10:
                row, col = _u16(payload, 0), _u16(payload, 2)
                sst_idx = _u32(payload, 6)
                value: Any = sst[sst_idx] if sst_idx < len(sst) else ""
            elif code == 0x0204 and len(payload) >= 8:
                row, col = _u16(payload, 0), _u16(payload, 2)
                value = _decode_short_biff_string(payload, 6, encoding)
            elif code == 0x0203 and len(payload) >= 14:
                row, col = _u16(payload, 0), _u16(payload, 2)
                value = struct.unpack_from("<d", payload, 6)[0]
            elif code == 0x0006 and len(payload) >= 14:
                row, col = _u16(payload, 0), _u16(payload, 2)
                value = "" if payload[6:8] == b"\xff\xff" else struct.unpack_from("<d", payload, 6)[0]
            elif code == 0x027E and len(payload) >= 10:
                row, col = _u16(payload, 0), _u16(payload, 2)
                value = _rk_value(_u32(payload, 6))
            elif code == 0x00BD and len(payload) >= 6:
                row, first_col, last_col = _u16(payload, 0), _u16(payload, 2), _u16(payload, len(payload) - 2)
                offset = 4
                for col in range(first_col, last_col + 1):
                    if offset + 6 <= len(payload):
                        cells.setdefault(row, {})[col] = _rk_value(_u32(payload, offset + 2))
                    offset += 6
                continue
            elif code == 0x00E5 and len(payload) >= 2:
                count = _u16(payload, 0)
                offset = 2
                for _ in range(count):
                    if offset + 8 <= len(payload):
                        merges.append((_u16(payload, offset), _u16(payload, offset + 2), _u16(payload, offset + 4), _u16(payload, offset + 6)))
                    offset += 8
                continue
            else:
                continue
            cells.setdefault(row, {})[col] = value

        for first_row, last_row, first_col, last_col in merges:
            value = cells.get(first_row, {}).get(first_col, "")
            if value in ("", None):
                continue
            for row in range(first_row, last_row + 1):
                for col in range(first_col, last_col + 1):
                    cells.setdefault(row, {}).setdefault(col, value)

        max_row = max(cells.keys(), default=-1)
        max_col = max((max(row_cells.keys()) for row_cells in cells.values() if row_cells), default=-1)
        rows = [[cells.get(row, {}).get(col, "") for col in range(max_col + 1)] for row in range(max_row + 1)]
        result.append((sheet_name, rows))
    return result


def parse_xls_rows(file_bytes: bytes) -> list[tuple[str, list[list[Any]]]]:
    try:
        import xlrd  # type: ignore[import-not-found]
    except ImportError as exc:
        try:
            return parse_xls_rows_without_xlrd(file_bytes)
        except Exception as fallback_exc:
            raise ValueError(".xls 손해보험 파일을 읽지 못했습니다. 파일을 다시 저장한 뒤 업로드해 주세요.") from fallback_exc

    workbook = xlrd.open_workbook(file_contents=file_bytes)
    result: list[tuple[str, list[list[Any]]]] = []
    for sheet in workbook.sheets():
        rows: list[list[Any]] = []
        for row_idx in range(sheet.nrows):
            rows.append([sheet.cell_value(row_idx, col_idx) for col_idx in range(sheet.ncols)])
        result.append((sheet.name, rows))
    return result


def parse_workbook(file_bytes: bytes, insurance_type: str = "life") -> list[dict[str, Any]]:
    if not file_bytes:
        raise ValueError("업로드된 파일이 비어 있습니다.")
    insurance_type = "nonlife" if insurance_type == "nonlife" else "life"
    if file_bytes.startswith(b"\xd0\xcf\x11\xe0\xa1\xb1\x1a\xe1"):
        if insurance_type != "nonlife":
            raise ValueError(".xls 파일은 손해보험 업로드에서만 지원합니다.")
        rows: list[dict[str, Any]] = []
        for sheet_name, values in parse_xls_rows(file_bytes):
            rows.extend(parse_nonlife_amount_sheet(values, sheet_name))
        return [normalize_rate_displays(row) for row in rows if row["company"] and row["product"] and not row_is_header_noise(row)]
    if not file_bytes.startswith(b"PK"):
        raise ValueError(".xlsx, .xlsm 또는 손해보험 .xls 파일만 업로드해 주세요.")
    workbook = load_workbook(BytesIO(file_bytes), data_only=True, read_only=False)
    rows: list[dict[str, Any]] = []
    for ws in workbook.worksheets:
        values = worksheet_rows(ws)
        if insurance_type == "nonlife":
            parsed_sheet = parse_nonlife_amount_sheet(values, ws.title)
            if not parsed_sheet:
                parsed_sheet.extend(parse_nonlife_sheet(values, ws.title))
        else:
            parsed_sheet = parse_life_sheet(values, ws.title)
            if not parsed_sheet:
                parsed_sheet.extend(parse_life_commission_rate_sheet(values, ws.title))
            if not parsed_sheet:
                parsed_sheet.extend(parse_life_company_detail_sheet(values, ws.title))
            if not parsed_sheet:
                parsed_sheet.extend(parse_monthly_matrix_sheet(values, ws.title))
            if not parsed_sheet:
                parsed_sheet.extend(parse_generic_rate_sheet(values, ws.title))
            for item in parsed_sheet:
                item.setdefault("insuranceType", "life")
                item.setdefault("metricMode", "percent")
        rows.extend(parsed_sheet)
    return [normalize_rate_displays(row) for row in rows if row["company"] and row["product"] and not row_is_header_noise(row)]


def parse_upload(body: bytes, content_type: str) -> tuple[bytes, dict[str, str]]:
    message = BytesParser(policy=policy.default).parsebytes(
        f"Content-Type: {content_type}\r\nMIME-Version: 1.0\r\n\r\n".encode("utf-8") + body
    )
    if not message.is_multipart():
        raise ValueError("multipart 형식이 아닙니다.")
    fields: dict[str, str] = {}
    file_bytes = b""
    for part in message.iter_parts():
        if part.get_filename():
            file_bytes = part.get_payload(decode=True) or b""
        else:
            name = part.get_param("name", header="content-disposition")
            if name:
                fields[str(name)] = clean(part.get_content())
    if not file_bytes:
        raise ValueError("업로드 파일을 찾지 못했습니다.")
    return file_bytes, fields


def session_from_cookie(cookie_header: str | None) -> str:
    if not cookie_header:
        return ""
    jar = cookies.SimpleCookie()
    jar.load(cookie_header)
    return jar.get("admin_session").value if "admin_session" in jar else ""


def make_session(role: str) -> tuple[str, dict[str, Any]]:
    token = secrets.token_urlsafe(32)
    session = {"expires": time.time() + SESSION_SECONDS, "csrf": secrets.token_urlsafe(32), "role": role}
    SESSIONS[token] = session
    return token, session


def valid_session(token: str) -> dict[str, Any] | None:
    if not token:
        return None
    session = SESSIONS.get(token)
    if not session:
        return None
    if session.get("expires", 0) < time.time():
        SESSIONS.pop(token, None)
        return None
    return session


def cleanup_state() -> None:
    now = time.time()
    for token, session in list(SESSIONS.items()):
        if session.get("expires", 0) < now:
            SESSIONS.pop(token, None)
    for ip, item in list(LOGIN_FAILURES.items()):
        if item.get("until", 0) and item.get("until", 0) < now:
            LOGIN_FAILURES.pop(ip, None)


def local_ip() -> str:
    try:
        sock = socket.socket(socket.AF_INET, socket.SOCK_DGRAM)
        sock.connect(("8.8.8.8", 80))
        ip = sock.getsockname()[0]
        sock.close()
        return ip
    except OSError:
        return "127.0.0.1"


class Handler(BaseHTTPRequestHandler):
    def end_headers(self) -> None:
        self.send_header("X-Content-Type-Options", "nosniff")
        self.send_header("X-Frame-Options", "DENY")
        self.send_header("Referrer-Policy", "no-referrer")
        self.send_header("Permissions-Policy", "camera=(), microphone=(), geolocation=()")
        self.send_header(
            "Content-Security-Policy",
            "default-src 'self'; img-src 'self' data:; style-src 'self' 'unsafe-inline'; "
            "script-src 'self' 'unsafe-inline'; object-src 'none'; base-uri 'self'; "
            "frame-ancestors 'none'; form-action 'self'",
        )
        self.send_header("Cache-Control", "no-store")
        super().end_headers()

    def send_json(self, status: int, payload: dict[str, Any], extra_headers: dict[str, str] | None = None) -> None:
        data = json.dumps(payload, ensure_ascii=False).encode("utf-8")
        self.send_response(status)
        self.send_header("Content-Type", "application/json; charset=utf-8")
        self.send_header("Content-Length", str(len(data)))
        if extra_headers:
            for key, value in extra_headers.items():
                self.send_header(key, value)
        self.end_headers()
        self.wfile.write(data)

    def redirect_home(self) -> None:
        self.send_response(303)
        self.send_header("Location", "/")
        self.end_headers()

    def read_json(self) -> dict[str, Any]:
        length = int(self.headers.get("Content-Length", "0"))
        if length <= 0:
            return {}
        if length > MAX_JSON_BYTES:
            raise ValueError("요청 본문이 너무 큽니다.")
        return json.loads(self.rfile.read(length).decode("utf-8"))

    def is_admin(self) -> bool:
        session = self.session()
        return bool(session and session.get("role") == "admin")

    def is_authenticated(self) -> bool:
        return bool(self.session())

    def session(self) -> dict[str, Any] | None:
        return valid_session(session_from_cookie(self.headers.get("Cookie")))

    def csrf_ok(self, submitted: str) -> bool:
        session = self.session()
        return bool(session and submitted and hmac.compare_digest(submitted, str(session.get("csrf", ""))))

    def client_ip(self) -> str:
        return self.client_address[0] if self.client_address else "unknown"

    def login_locked(self) -> bool:
        item = LOGIN_FAILURES.get(self.client_ip())
        return bool(item and item.get("count", 0) >= MAX_LOGIN_FAILURES and item.get("until", 0) > time.time())

    def record_login_failure(self) -> None:
        item = LOGIN_FAILURES.setdefault(self.client_ip(), {"count": 0, "until": 0})
        item["count"] = int(item.get("count", 0)) + 1
        if item["count"] >= MAX_LOGIN_FAILURES:
            item["until"] = time.time() + LOCK_SECONDS

    def clear_login_failure(self) -> None:
        LOGIN_FAILURES.pop(self.client_ip(), None)

    def session_cookie(self, token: str, max_age: int = SESSION_SECONDS) -> str:
        forwarded_proto = self.headers.get("X-Forwarded-Proto", "").split(",")[0].strip().lower()
        secure = "; Secure" if forwarded_proto == "https" else ""
        return f"admin_session={token}; Path=/; Max-Age={max_age}; HttpOnly; SameSite=Lax{secure}"

    def read_upload_body(self) -> tuple[bytes, str]:
        length = int(self.headers.get("Content-Length", "0"))
        if length <= 0:
            raise ValueError("업로드된 파일을 찾지 못했습니다.")
        if length > MAX_UPLOAD_BYTES:
            raise ValueError("업로드 파일이 너무 큽니다. 20MB 이하 파일만 업로드해 주세요.")
        return self.rfile.read(length), self.headers.get("Content-Type", "")

    def do_GET(self) -> None:
        cleanup_state()
        if self.path == "/api/session":
            session = self.session()
            self.send_json(
                200,
                {
                    "authenticated": bool(session),
                    "admin": bool(session and session.get("role") == "admin"),
                    "role": session.get("role", "") if session else "",
                    "csrf": session.get("csrf", "") if session else "",
                },
            )
            return

        if self.path == "/api/data":
            if not self.is_authenticated():
                self.send_json(403, {"error": "로그인 후 조회할 수 있습니다."})
                return
            self.send_json(200, CURRENT_UPLOAD)
            return

        if self.path.startswith("/assets/"):
            target = (ROOT / self.path.lstrip("/")).resolve()
            assets_root = (ROOT / "assets").resolve()
            try:
                target.relative_to(assets_root)
            except ValueError:
                self.send_json(404, {"error": "Not found"})
                return
            if not target.exists() or not target.is_file():
                self.send_json(404, {"error": "Not found"})
                return
            data = target.read_bytes()
            content_type = mimetypes.guess_type(target.name)[0] or "application/octet-stream"
            self.send_response(200)
            self.send_header("Content-Type", content_type)
            self.send_header("Content-Length", str(len(data)))
            self.end_headers()
            self.wfile.write(data)
            return

        html = (ROOT / "index.html").read_text(encoding="utf-8")
        session = self.session()
        visible_upload = empty_upload(message="로그인 후 조회할 수 있습니다.")
        if session:
            visible_upload = CURRENT_UPLOAD
        payload = json.dumps(visible_upload, ensure_ascii=False).replace("</", "<\\/")
        initial_data = f"<script>window.__INITIAL_UPLOAD__ = {payload};</script>\n"
        app_script_marker = "<script>\n    const state"
        if app_script_marker in html:
            html = html.replace(app_script_marker, initial_data + app_script_marker, 1)
        else:
            html = html.replace("</body>", initial_data + "</body>")
        data = html.encode("utf-8")
        self.send_response(200)
        self.send_header("Content-Type", "text/html; charset=utf-8")
        self.send_header("Content-Length", str(len(data)))
        self.end_headers()
        self.wfile.write(data)

    def do_POST(self) -> None:
        global CURRENT_UPLOAD
        cleanup_state()

        if self.path == "/upload":
            session = self.session()
            if not session or session.get("role") != "admin":
                self.redirect_home()
                return
            try:
                body, content_type = self.read_upload_body()
                file_bytes, fields = parse_upload(body, content_type)
                if not hmac.compare_digest(fields.get("csrf", ""), str(session.get("csrf", ""))):
                    self.redirect_home()
                    return
                rows = parse_workbook(file_bytes, fields.get("insuranceType", "life"))
                existing_rows = CURRENT_UPLOAD.get("rows", [])
                combined_rows = merge_uploaded_rows(existing_rows, rows)
                CURRENT_UPLOAD = {
                    "rows": combined_rows,
                    "fileName": "업로드된 엑셀 파일",
                    "message": "" if rows else "인식 가능한 수수료 표를 찾지 못했습니다.",
                    "error": "",
                }
                save_upload(CURRENT_UPLOAD)
            except Exception as exc:
                CURRENT_UPLOAD = {"rows": [], "fileName": "", "message": "", "error": f"엑셀 파일을 읽지 못했습니다: {exc}"}
            self.redirect_home()
            return

        if self.path == "/api/login":
            try:
                if self.login_locked():
                    self.send_json(429, {"error": "로그인 시도가 너무 많습니다. 잠시 후 다시 시도해 주세요."})
                    return
                payload = self.read_json()
                login_id = str(payload.get("id", "")).strip().lower()
                password = str(payload.get("password", ""))
                role = ""
                if login_id in ADMIN_LOGIN_IDS and hmac.compare_digest(password, ADMIN_PASSWORD):
                    role = "admin"
                elif login_id in VIEWER_LOGIN_IDS and hmac.compare_digest(password, VIEWER_PASSWORD):
                    role = "viewer"

                if role:
                    self.clear_login_failure()
                    token, session = make_session(role)
                    self.send_json(
                        200,
                        {"authenticated": True, "admin": role == "admin", "role": role, "csrf": session["csrf"]},
                        {"Set-Cookie": self.session_cookie(token)},
                    )
                    return
                self.record_login_failure()
                self.send_json(401, {"error": "아이디 또는 비밀번호가 올바르지 않습니다."})
            except Exception as exc:
                self.send_json(400, {"error": str(exc)})
            return

        if self.path == "/api/logout":
            token = session_from_cookie(self.headers.get("Cookie"))
            SESSIONS.pop(token, None)
            self.send_json(200, {"admin": False}, {"Set-Cookie": "admin_session=; Path=/; Max-Age=0; HttpOnly; SameSite=Lax"})
            return

        if self.path == "/api/switch-viewer":
            session = self.session()
            if not session or session.get("role") != "admin":
                self.send_json(403, {"error": "관리자 로그인 후 조회 모드로 전환할 수 있습니다."})
                return
            if not hmac.compare_digest(self.headers.get("X-CSRF-Token", ""), str(session.get("csrf", ""))):
                self.send_json(403, {"error": "보안 토큰이 올바르지 않습니다. 새로고침 후 다시 시도해 주세요."})
                return
            session["role"] = "viewer"
            session["csrf"] = secrets.token_urlsafe(32)
            self.send_json(200, {"authenticated": True, "admin": False, "role": "viewer", "csrf": session["csrf"]})
            return

        if self.path == "/api/upload":
            session = self.session()
            if not session or session.get("role") != "admin":
                self.send_json(403, {"error": "관리자 로그인 후 업로드해 주세요."})
                return
            if not hmac.compare_digest(self.headers.get("X-CSRF-Token", ""), str(session.get("csrf", ""))):
                self.send_json(403, {"error": "보안 토큰이 올바르지 않습니다. 새로고침 후 다시 시도해 주세요."})
                return
            try:
                body, content_type = self.read_upload_body()
                file_bytes, fields = parse_upload(body, content_type)
                rows = parse_workbook(file_bytes, fields.get("insuranceType", "life"))
                message = "" if rows else "인식 가능한 수수료 표를 찾지 못했습니다."
                existing_rows = CURRENT_UPLOAD.get("rows", [])
                combined_rows = merge_uploaded_rows(existing_rows, rows)
                CURRENT_UPLOAD = {
                    "rows": combined_rows,
                    "fileName": "업로드된 엑셀 파일",
                    "message": message,
                    "error": "",
                }
                save_upload(CURRENT_UPLOAD)
                self.send_json(200, {**CURRENT_UPLOAD, "addedRows": len(rows)})
            except Exception as exc:
                self.send_json(400, {"error": f"엑셀 파일을 읽지 못했습니다: {exc}"})
            return

        if self.path == "/api/clear-upload":
            session = self.session()
            if not session or session.get("role") != "admin":
                self.send_json(403, {"error": "관리자 로그인 후 삭제할 수 있습니다."})
                return
            if not hmac.compare_digest(self.headers.get("X-CSRF-Token", ""), str(session.get("csrf", ""))):
                self.send_json(403, {"error": "보안 토큰이 올바르지 않습니다. 새로고침 후 다시 시도해 주세요."})
                return
            payload = self.read_json()
            source = str(payload.get("source", "")).strip()
            if source:
                rows = [row for row in CURRENT_UPLOAD.get("rows", []) if str(row.get("source", "")) != source]
                removed_count = len(CURRENT_UPLOAD.get("rows", [])) - len(rows)
                CURRENT_UPLOAD = {
                    "rows": rows,
                    "fileName": CURRENT_UPLOAD.get("fileName", "업로드된 엑셀 파일") if rows else "",
                    "message": "",
                    "error": "",
                }
                save_upload(CURRENT_UPLOAD)
                self.send_json(200, {**CURRENT_UPLOAD, "message": f"{source} 구분 데이터가 삭제되었습니다.", "removedRows": removed_count})
                return
            CURRENT_UPLOAD = empty_upload()
            save_upload(CURRENT_UPLOAD)
            self.send_json(200, {**CURRENT_UPLOAD, "message": "업로드된 엑셀 데이터가 전체 삭제되었습니다."})
            return

        self.send_json(404, {"error": "Not found"})

    def log_message(self, format: str, *args: Any) -> None:
        return


if __name__ == "__main__":
    host = "0.0.0.0"
    port = int(os.environ.get("PORT", "8766"))
    server = ThreadingHTTPServer((host, port), Handler)
    print(f"Local:   http://127.0.0.1:{port}/")
    print(f"Network: http://{local_ip()}:{port}/")
    server.serve_forever()
